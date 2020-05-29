#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Part of XLSX to Odoo import
# Copyright 2020 David Todd <dtodd@oceantech.com>
# License: MIT License, refer to `license.md` for more information

# pylint: disable=import-error

"""
    A Really hacky script (not really), who's purpose
    is to import oddly formatted data that was provided
    on a customer pickup to data that the erp is
    expecting. This will accomplish the import via
    the XMLRPC interface
"""

import os
import csv
import time
import logging
import itertools

from openpyxl import load_workbook
from record import Record
from api import API

# Odoo Stuff
# The database ID of the asset catalog we are importing into
ASSET_CATALOG_ID = int(os.environ.get('odoo_asset_catalog_id', 0))
# The database ID of the data destruction we are importing into
DATA_DESTRUCTION_ID = int(os.environ.get('odoo_data_destruction_id', 0))

# Spreadsheet Stuff
SPREADSHEET = os.environ.get('spreadsheet', '')
SHEET = os.environ.get('sheet', '')
FIRST_ROW = int(os.environ.get('first_row', 1))
LAST_ROW = int(os.environ.get('last_row', 2000))
LAST_COL = int(os.environ.get('last_col', 6))

# Items in this list will always create a Record (though that record
# won't get uploaded to the ERP), and will additionally be added to
# a spreadsheet as they are come across
SERIALS_TO_IGNORE = os.environ.get('serials_to_ignore').strip().split('\n')

FILENAME_TIME = '%s' % (time.time())
IGNORE_CSV = '%s.csv' % (FILENAME_TIME)

# Logging
logging.basicConfig(
    filename='%s.log' % (FILENAME_TIME),
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%m/%d/%Y %I:%M:%S %p'
)
# Log to console and file
CONSOLE = logging.StreamHandler()
CONSOLE.setLevel(logging.INFO)
FORMATTER = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
CONSOLE.setFormatter(FORMATTER)
logging.getLogger('').addHandler(CONSOLE)

# Sanity Checks
if ASSET_CATALOG_ID <= 0:
    logging.warning('Zero or negative asset catalog. Records will not get uploaded to it')
if DATA_DESTRUCTION_ID <= 0:
    logging.warning('Zero or negative data destruction. Records will not get uploaded to it')

class ProcessWorkbook:
    """
        Provides a mechanism for extracting the content
        from the workbook and uploading it to Odoo
    """

    # pylint: disable=too-many-instance-attributes
    def __init__(self):
        self.api = API()
        self.workbook = load_workbook(filename=SPREADSHEET, data_only=True)[SHEET]

        # Special Serials
        self.ignore_csv_file = open(IGNORE_CSV, 'w')
        self.ignore_csv = csv.DictWriter(
            self.ignore_csv_file,
            fieldnames=[
                'serial', 'asset_tag', 'make',
                'model', 'device_type', 'children'
            ],
            dialect=csv.excel
        )
        self.ignore_csv.writeheader()

        # Failed records are rows that for one reason or another didn't generate a Record object
        self.failed_records = list()
        self.records = list()
        self.records_to_upload = list()
        self.last_parent = None

        # These lists are expected to contain one or more Tuples
        # Each Tuple will be of the format (make, model)
        self.models_to_search = list()
        self.models_to_create = list()
        # The Tuples here will be of the format ((make, model), id)
        self.models_to_ids = list()

        # Serials that match items in this list will always
        # be returned False from `self.serial_in_records`
        self.serials_to_ignore = [None, '', 'N/A']
        self.serials_to_ignore.extend(SERIALS_TO_IGNORE)

        self.rows_processed = 0
        self.sorting_records_uploaded = 0
        self.data_records_uploaded = 0
        self.records_ignored = 0

        logging.info('Initialized ProcessWorkbook')

    def __del__(self):
        """
            Automatically closes file handlers when destructed normally
        """
        self.ignore_csv_file.close()

        logging.info('Processed %d rows', (self.rows_processed))
        logging.info('Created %d Records', (len(self.records)))
        logging.info('Uploaded %d Sorting Assets', (self.sorting_records_uploaded))
        logging.info('Uploaded %d Data Destruction Assets', (self.data_records_uploaded))
        logging.info('Prevented %d Records from being uploaded', (self.records_ignored))

        for row in self.failed_records:
            # pylint: disable=logging-not-lazy
            logging.info(
                'Row that failed Record Creation: %s | %s | %s | %s | %s' % (
                    row[0].value,
                    row[1].value,
                    row[3].value,
                    row[4].value,
                    row[5].value
                )
            )

        logging.info('ProcessWorkbook Finished')

    def get_id_from_model(self, model):
        """
            Searches `self.models_to_ids` for
            a matching `model`, and returns the
            id if there is one. Returns None
            otherwise
        """
        for item in self.models_to_ids:
            if model == item[0][1]:
                return item[1]
        return None

    def serial_in_records(self, serial, records=None):
        """
            Determines if a particular serial number
            is in a record list, and returns True if so;
            otherwise, returns False

            `serial` is a string to check against the
            record list

            `records` is an optional argument that will
            search that specific record list (such as a
            parent's children) instead of just the parent
            records.

            If the serial is to be ignored, this will
            always return False, to allow the Record object
            to be created (used to save it to a csv)
        """
        if records is None:
            records = self.records
        if serial in self.serials_to_ignore:
            return False
        if serial in [record.serial for record in records if record.serial]:
            return True
        return False

    def create_record_from_row(self, row, parent=True, search_model=True):
        """
            With a provided `row`, this method will first search
            for a matching serial number and if it doesn't exist,
            a new Record will be created with that row's data.
            If the row's serial number is special, the record will
            be created regardless of one already existing.

            `parent` is an optional argument that sets the `last_parent`
            attribute. `last_parent` is used to append new Records to
            a parent's `children` attribute.

            `search_model` is an optional argument that when True
            will search Odoo for that record's model (to be created
            if it can't be found)

            If the serial number isn't in the list, this returns the
            created Record object. Otherwise, it returns False
        """

        serial = str(row[0].value)
        if not self.serial_in_records(serial):
            # pylint: disable=bad-whitespace
            logging.debug('Creating Record for %s', (serial))
            record = Record(
                serial = serial,
                asset_tag = str(row[1].value),
                make = str(row[3].value),
                model = str(row[4].value),
                device_type = str(row[5].value),
                children = None
            )

            if parent:
                record.children = list()
                self.last_parent = record

            if search_model:
                # Add unique models so we can search for their sellable ids
                if record.model not in itertools.chain(*self.models_to_search):
                    self.models_to_search.append((record.make, record.model))

            return record
        return False

    def build_record_list(self):
        """
            Reads the workbook and sorts each row into
            multiple instances of the Record class.

            When this method comes across a "Parent"
            record, it stores that Record instance, which
            will be used whenever this method comes across
            a "Child" record.

            When this method comes across a "Child"
            record, it gets appended to the `children`
            list of the stored parent object.

            When there is no relationship, the Record
            will still be created, without any Children

            Returns `self` (this instance of ProcessWorkbook)
        """
        logging.info('Getting rows from the spreadsheet and sorting relationships')
        # pylint: disable=bad-continuation
        for row in self.workbook.iter_rows(
            min_row=FIRST_ROW,
            max_col=LAST_COL,
            max_row=LAST_ROW
        ):

            relationship = row[2].value

            if relationship == 'Parent':
                record = self.create_record_from_row(row, True)
                if record:
                    self.records.append(record)

            elif relationship == 'Child':
                record = self.create_record_from_row(row, False, False)
                if record:
                    self.last_parent.children.append(record)

            else:
                record = self.create_record_from_row(row, False)
                if record:
                    self.records.append(record)

            if not record:
                self.failed_records.append(row)

            self.rows_processed += 1

        return self

    def get_records(self):
        """
            Deprecated, use `build_record_list` instead.
        """
        logging.warning('get_records is deprecated and will be removed in the future.')
        logging.warning('Please use `build_record_list` instead.')
        return self.build_record_list()

    def show_records(self):
        """
            Logs all of the records stored, in JSON format
        """
        # pylint: disable=unnecessary-comprehension
        logging.debug([record for record in self.records])

    def get_odoo_model_ids(self):
        """
            Iterates over unique models and searches
            Odoo for the database id of those models.
            Once the search is complete, `self.models_to_ids`
            contains a mapping between each unique model
            name and the database id.

            When a model can't be found, that model is logged
            so that a manual search can be done, or a new item
            can be created.

            When a model returns multiple ids, that model is
            logged so that a manual search can be done to
            select the "correct" database id.

            Returns `self` (this instance of ProcessWorkbook)
        """
        logging.info('Searching Odoo for sellable items with matching models')
        for model in self.models_to_search:
            odoo_records = self.api.do_search_and_read(
                'erpwarehouse.sellable',
                [('model', 'ilike', model[1])]
            )

            if not odoo_records:
                logging.warning('Unable to find model: %s', (model[1]))
                self.models_to_create.append(model)
            else:
                # If the search returns more than one, we'll
                # assume that it was the first one since some
                # models are duplicated (for whatever reason)
                self.models_to_ids.append((model, odoo_records[0]['id']))

    def create_missing_model_ids(self):
        """
            For any models that couldn't be located
            when initally searched, create them and
            then store those newly created ids

            Returns `self` (this instance of ProcessWorkbook)
        """
        logging.info('Creating sellable items for missing models')
        for model in self.models_to_create:
            logging.info('Creating model: %s', (model[1]))
            result = self.api.do_create(
                'erpwarehouse.sellable',
                {
                    'make': model[0],
                    'model': model[1]
                }
            )
            self.models_to_ids.append((model, result))

        return self

    def asset_line_exists(self, record):
        """
            Searches the asset catalog for
            a line item matching the provided `record`.

            Determines if a line is the same if either the
            serial number was previously recorded.

            Returns True if there is an existing record in Odoo,
            False otherwise.
        """
        logging.debug('Checking if "%s" already exists before creation in Odoo', (record.serial))
        result = self.api.do_search(
            'erpwarehouse.asset',
            [
                ('catalog', '=', ASSET_CATALOG_ID),
                ('make', '=', self.get_id_from_model(record.model)),
                ('serial', '=ilike', record.serial),
            ]
        )
        if len(result) > 0:
            return True
        return False

    def _create_asset_catalog_line(self, record):
        """
            With the provided `record` (Record) instance,
            this method will issue an API request to Odoo
            to create the line item after searching Odoo
            for that record.
        """
        if self.get_id_from_model(record.model):
            if not self.asset_line_exists(record):
                result = self.api.do_create(
                    'erpwarehouse.asset',
                    {
                        'catalog': ASSET_CATALOG_ID,
                        'make': self.get_id_from_model(record.model),
                        'serial': record.serial,
                        'tag': record.asset_tag,
                    }
                )
                self.sorting_records_uploaded += 1
                logging.debug('Added id: %s', (result))
            else:
                logging.warning('"%s" already existed, so it was skipped', (record.serial))
        else:
            logging.error('Unable to add "%s" as there is no sellable id', (record.serial))

        return self

    def _create_data_destruction_line(self, record, child=False):
        """
            With the provided `record` and optional `child` Record
            instances, this method will issue an API request to
            Odoo to create the line item
        """

        device_type = '0'
        if record.device_type == 'Hard Drive':
            device_type = 'H'
        elif record.device_type == 'Network':
            device_type = 'N'
        elif record.device_type == 'Tape':
            device_type = 'T'

        if child:
            if child.device_type == 'Hard Drive':
                device_type = 'H'
            elif child.device_type == 'Network':
                device_type = 'N'
            elif child.device_type == 'Tape':
                device_type = 'T'

        if self.get_id_from_model(record.model):
            result = self.api.do_create(
                'erpwarehouse.ddl_item',
                {
                    'ddl': DATA_DESTRUCTION_ID,
                    'make': self.get_id_from_model(record.model),
                    'serial': record.serial,
                    'storser': child.serial if child else 'N/A',
                    'type': device_type,
                })
            self.data_records_uploaded += 1
            logging.debug('Added id: %s', (result))
        else:
            logging.error('Unable to add "%s" as there is no sellable id', (record.serial))

        return self

    def create_line_items(self):
        """
            For all the records that we can
            process (there's a mapped model),
            create the asset catalog and
            data destruction line items.

            Returns `self` (this instance of ProcessWorkbook)
        """
        logging.info('Creating Line items for accepted records in Odoo')
        for record in self.records_to_upload:

            if ASSET_CATALOG_ID:
                self._create_asset_catalog_line(record)

            if DATA_DESTRUCTION_ID:
                if not record.children:
                    self._create_data_destruction_line(record)
                else:
                    for child in record.children:
                        self._create_data_destruction_line(record, child)

        return self

    def remove_ignored_records(self):
        """
            Populates `self.records_to_upload` with
            any record that is not also a part of the
            `self.serials_to_ignore` list, counts the
            serials that were ignored, and writes those
            rows to an ignore csv
        """
        logging.info('Removing Ignored Serials from Records')

        for record in self.records:
            if record.serial in self.serials_to_ignore:
                self.records_ignored += 1
                logging.warning(
                    '"%s" is special, skipping import and saving to special list', (record.serial)
                )
                self.ignore_csv.writerow({
                    'serial': record.serial,
                    'asset_tag': record.asset_tag,
                    'make': record.make,
                    'model': record.model,
                    'device_type': record.device_type,
                    'children': record.children,
                })
            else:
                self.records_to_upload.append(record)

    def run(self):
        """
            Runs everything in the order that is required
        """
        self.build_record_list()
        self.show_records()
        self.get_odoo_model_ids()
        self.create_missing_model_ids()
        self.remove_ignored_records()
        self.create_line_items()

if __name__ == '__main__':
    ProcessWorkbook().run()
